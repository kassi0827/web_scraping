import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime
import xlwings as xw

# 更新対象ファイルのパス
excel_path = '/Users/kashimataichi/play_python/web_scraping/stock_price/vt/vt_data.xlsx'


# ファイルのデータがある最終行と最終列を取得
wb = openpyxl.load_workbook(excel_path)
ws = wb['vt_data']
max_row_plus_one = ws.max_row + 1


# エクセル転記処理用の連想配列
dict_excel_input = {
    'B列始値': (2, 0),
    'C列前日終値': (3, 3),
    'D列出来高': (4, 2),
    'E列1年TR': (5, 22),
    'F列3年TR': (6, 19),
    'G列5年TR': (7, 20),
    'H列年初来R': (8, 23),
    'I列日次安値-高値レンジ': (9, 1),
    'J列52週レンジ': (10, 4),
    'K列資産総額': (11, 11),
}


# 年月日の転記
dt_now = datetime.datetime.now()
today_yyyy_mm_dd = str(dt_now.year) + '/' + \
    str(dt_now.month) + '/' + str(dt_now.day)
ws.cell(row=max_row_plus_one, column=1).value = today_yyyy_mm_dd


# 対象ページのソースを取得
vt_bloomberg_url = 'https://www.bloomberg.co.jp/quote/VT:US'
res = requests.get(vt_bloomberg_url)

# 取得したページのソースをビューティフルスープでスクレイピング
# 転記の処理はfor文化を検討
soup = BeautifulSoup(res.text, 'html.parser')
elems = soup.find_all('div', attrs={'class': 'cell__value cell__value_'})

elems_others = soup.find_all(
    'div', attrs={'class': 'cell__value cell__value_down'})

for element in elems_others:
    elems.append(element)

# 実際の転記処理
for execution in dict_excel_input:
    ws.cell(row=max_row_plus_one,
            column=dict_excel_input[execution][0]).value = elems[dict_excel_input[execution][1]].contents[0]


wb.save(excel_path)

xw.Book(r'/Users/kashimataichi/play_python/web_scraping/stock_price/vt/vt_data.xlsx')

print('Transaction completed.')
